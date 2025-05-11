from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from math import sqrt
import os
from csv import reader

path = 'C:\\Users\\ferna\\Downloads\\admin_inventario.xlsx'


class program:
    def __init__(self, path):
      os.system('cls')  # Limpia la consola
      self.path = path
      self.wb = load_workbook(path)
      self.ws = self.wb['inventario']
      self.data = {}
      self.procedimiento()
      self.demanda_anual()
      self.costo_total()
      self.costo_unitario()
      self.metodo_eoq()
      self.punto_de_reorden()
      self.clasificacion()
      self.grabar()
      self.ajustar_ancho_columnas()
      self.wb.save(self.path)
      self.wb.close()
    
    def procedimiento(self):
      # self.data = self.inventario_original
      for row in self.ws.iter_rows(min_row=9, min_col=2, values_only=True):  
        self.data[row[0]] = {
          'Clave' : row[1], 
          'Articulo' : row[2],
          'Descripcion' : row[3], 
          'Demanda diaria' : row[4],
          'Demanda anual' : 0,
          'Costo x pedido' : row[6],
          'Costo x mantenimiento' : row[7],
          'Tiempo de entrega' : row[8],
          'Dias stock' : row[9],
          'ene-24' : row[10],
          'feb-24' : row[11],
          'mar-24' : row[12],
          'abr-24' : row[13],
          'may-24' : row[14],
          'jun-24' : row[15],
          'Costo total' : 0, 
          'Clasificacion' : 'x', 
          'Costo unitario' : 0,
          'EOQ': 0,
          'Punto de reorden' : 0
        }

    def demanda_anual(self):
      for key, value in self.data.items():
        self.data[key]['Demanda anual'] = value['Demanda diaria'] * 365

    def costo_total(self):
      for key, value in self.data.items():
        self.data[key]['Costo total'] = value['ene-24'] + value['feb-24'] + value['mar-24'] + value['abr-24'] + value['may-24'] + value['jun-24']

    def costo_unitario(self):
      for key, value in self.data.items():
        self.data[key]['Costo unitario'] = value['Costo total'] / 6

    def metodo_eoq(self):
      for key, value in self.data.items():
        self.data[key]['EOQ'] = sqrt((2 * value['Costo x pedido'] * value['Demanda anual']) / (value['Costo x mantenimiento']))

    def punto_de_reorden(self):
      for key, value in self.data.items():
        self.data[key]['Punto de reorden'] = (value['Demanda diaria'] * value['Tiempo de entrega']) + (value['Dias stock'] * value['Demanda diaria'])

    def clasificacion(self):
      self.sorted_data = dict(sorted(self.data.items(), key=lambda item: item[1]['Costo total'], reverse=True))
      while True:
          var_a = float(input('\aIngrese el porcentaje de A: ')) / 100 
          var_b = float(input('Ingrese el porcentaje de B: ')) / 100 
          var_c = float(input('Ingrese el porcentaje de C: ')) / 100 
          
          if round(var_a + var_b + var_c, 4) == 1.0:  # Se asegura de que la suma sea exactamente 100%
              total_items = len(self.sorted_data)
              limite_a = round(total_items * var_a)
              limite_b = round(total_items * var_b) + limite_a
              limite_c = total_items  # El resto de los elementos
              
              keys = list(self.sorted_data.keys())
              
              for i in range(limite_a):
                  self.sorted_data[keys[i]]['Clasificacion'] = 'A'
              for i in range(limite_a, limite_b):
                  self.sorted_data[keys[i]]['Clasificacion'] = 'B'
              for i in range(limite_b, limite_c):
                  self.sorted_data[keys[i]]['Clasificacion'] = 'C'
              break
          else:
              print('Los porcentajes no suman 100%, intente nuevamente.\n')

    
    def grabar(self):
      # escribe los encabezados
      encabezados = [
        'ID', 'Clave', 'Articulo', 'Descripcion', 'Demanda diaria',
        'Demanda anual', 'Costo x pedido', 'Costo x mantenimiento', 'Tiempo de entrega',
        'Dias stock', 'ene-24', 'feb-24', 'mar-24', 'abr-24', 'may-24', 'jun-24',
        'Costo total', 'Clasificacion', 'Costo unitario', 'EOQ', 'Punto de reorden'
        ]
      col_inicial_encabezados = 26
      for i in encabezados:
        col_inicial_encabezados += 1
        cell = self.ws.cell(row=8, column=col_inicial_encabezados, value=i)
        cell.alignment = Alignment(horizontal='center')
      
      # escribe los ids
      col_inicial_datos = 27
      fila_inicial_ids = 8
      for key in self.sorted_data:
        fila_inicial_ids += 1
        cell = self.ws.cell(row=fila_inicial_ids, column=col_inicial_datos, value=key)
        cell.alignment = Alignment(horizontal='center')
      
      # escribe los datos
      fila_inicial = 9  # Empezamos en la fila B9
      for fila, (key, value) in enumerate(self.sorted_data.items(), start=fila_inicial):
          for col, atributo in enumerate(value.values(), start=col_inicial_datos+1):
              cell = self.ws.cell(row=fila, column=col, value=atributo)
              cell.alignment = Alignment(horizontal='center')
      
      self.ws.merge_cells(start_row=6, start_column=27, end_row=6, end_column=48)
      cell = self.ws.cell(row=6, column=27, value='INVENTARIO CALCULADO')
      cell.alignment = Alignment(horizontal='center')
      cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
      
      for row in range(9, 101):
        cell = self.ws.cell(row=row, column=29)
        cell.alignment = Alignment(horizontal='left')
        
      for row in range(9, 101):
        cell = self.ws.cell(row=row, column=30)
        cell.alignment = Alignment(horizontal='left')

    def ajustar_ancho_columnas(self):
      # Ajustar el ancho de las columnas desde AA (columna 27) hasta AV (columna 48)
      for col in range(27, 49):  # Columnas AA (27) a AV (48)
        max_length = 0
        column_letter = self.ws.cell(row=1, column=col).column_letter  # Obtener la letra de la columna
        for row in self.ws.iter_rows(min_row=6, max_row=self.ws.max_row, min_col=col, max_col=col):
          for cell in row:
              if cell.value:  # Si la celda tiene un valor
                  max_length = max(max_length, len(str(cell.value)))
        # Ajustar el ancho de la columna
        self.ws.column_dimensions[column_letter].width = max_length + 2  # Agregar un pequeño margen


mi_programa = program(path)